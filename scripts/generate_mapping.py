"""
ConformEDI - Génération de la table de correspondance enrichie
Lit l'Excel original et ajoute les colonnes de mapping vers les variables diagnostic.
"""
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SOURCE = '/mnt/user-data/uploads/2025-12-24-Questionnaires_2026_CLEAN.xlsx'
DEST = '/home/claude/work/conformedi_mapping_a_valider.xlsx'

SHEETS_WITH_QUESTIONS = [
    'Socle règlementaire',
    "Système d'information",
    "Délégation d'encaissement-décai",
    'Traitement donnée de santé',
    'Subdélégation',
    'Socle distributeur',
    'Entrée en relation',
]


def classify(ref, theme, question, type_reponse):
    """Retourne (bloc_source, diagnostic_var, conversion_rule, confidence, notes)"""
    q = (question or "").lower()
    t = (theme or "").lower()
    ref = (ref or "").strip()

    # === Documents et pièces justificatives ===
    doc_keywords = ['kbis', 'orias', 'iban', 'statuts', 'cerfa', 'pièce d\'identité',
                    'cv', 'liasses fiscales', 'rapport des commissaires', 'bilan']
    if any(k in q for k in doc_keywords):
        return ('PJ_DOC', None, 'OUI→Oui, NON→Non',
                'high', 'Question de pièce justificative — réponse = disponibilité du document')

    # === Champs numériques ou texte libre ===
    if type_reponse in ('Numérique', 'Pourcentage', 'Enumération', 'Coordonnées'):
        return ('INPUT_DIRECT', None, 'valeur brute',
                'high', f'Champ {type_reponse} — à extraire directement, pas de mapping diagnostic')

    # === Activités hors France / EEE / opérations ===
    if 'exclusivement traitées' in q and 'france' in q:
        return ('bloc_1', 'activites_operationnelles_traitees_hors_france',
                'OUI→Non, NON→Oui',
                'high', 'Logique INVERSÉE : question "exclusivement en France ?"')
    if 'implantations hors eee' in q or 'hors eee' in q and 'filiales' in q:
        return ('bloc_1', 'activite_hors_eee', 'OUI→Oui, NON→Non',
                'high', '')
    if 'sous-traitant' in q and 'union européenne' in q:
        return ('bloc_1', 'activites_operationnelles_traitees_hors_france',
                'OUI→Oui, NON→Non',
                'medium', 'Sous-traitance hors UE — proche de activites_operationnelles_hors_france')
    if 'autre état membre de l\'eee' in q or 'autre état de l\'eee' in q:
        return ('bloc_1', 'activite_eee', 'OUI→Oui, NON→Non',
                'high', '')
    if ('libre prestation' in q or 'libre établissement' in q or
        ('hors de france' in q and 'eee' in q)):
        return ('bloc_1', 'activite_eee', 'OUI→Oui, NON→Non',
                'medium', 'LPS/LE dans EEE')
    if 'courtage hors eee' in q or 'hors eee' in q and 'filiale' in q:
        return ('bloc_1', 'activite_hors_eee', 'OUI→Oui, NON→Non',
                'high', '')
    if 'mandataires d\'intermédiaire' in q or 'mia' in q:
        return ('INPUT_DIRECT', None, 'valeur brute',
                'high', 'MIA — donnée directe (oui/non + nombre)')

    # === LCB-FT / Sanctions ===
    if 'lcb' in t.lower() or 'blanchiment' in t.lower() or 'lcb' in q or 'tracfin' in q:
        if 'filtrage' in q or 'outil' in q and ('automatique' in q or 'automatisé' in q):
            return ('bloc_4', 'filtrage_lcbft', 'OUI→Oui, NON→Non',
                    'high', '')
        if 'mise à jour' in q and ('liste' in q or 'sanctions' in q):
            return ('bloc_4', 'mise_a_jour_sanctions', 'OUI→Oui, NON→Non',
                    'high', '')
        if 'régimes de sanctions' in q or 'onu' in q or 'régime' in q:
            return ('bloc_1', 'detection_ppe_sanctions', 'OUI→Oui, NON→Non',
                    'medium', 'Régimes de sanctions couverts (ONU/UE/FR/USA/UK)')
        if 'formation' in q:
            return ('bloc_1', 'formation_lcbft', 'OUI→Oui, NON→Non',
                    'high', '')
        if 'ppe' in q or 'politiquement exposée' in q or 'gel des avoirs' in q:
            return ('bloc_1', 'detection_ppe_sanctions', 'OUI→Oui, NON→Non',
                    'high', '')
        if 'procédure d\'alerte' in q or 'alerte' in q and 'assureur' in q:
            return ('bloc_1', 'detection_ppe_sanctions', 'OUI→Oui, NON→Non',
                    'medium', 'Procédure d\'alerte à l\'assureur')
        if 'responsable' in q or 'mise en œuvre' in q or 'désigné' in q:
            return ('bloc_1', 'lcbft', 'OUI→Oui, NON→Non',
                    'medium', 'Responsable LCB-FT')
        return ('bloc_1', 'lcbft', 'OUI→Oui, NON→Non',
                'medium', 'LCB-FT générique')

    # === Sanctions internationales ===
    if 'sanctions internationales' in t.lower():
        if 'conservez-vous' in q or 'vérifications' in q:
            return ('bloc_1', 'detection_ppe_sanctions', 'OUI→Oui, NON→Non',
                    'medium', 'Conservation vérifications sanctions')
        if 'contrôles' in q:
            return ('bloc_1', 'detection_ppe_sanctions', 'OUI→Oui, NON→Non',
                    'medium', 'Contrôles dispositif sanctions')
        return ('bloc_1', 'detection_ppe_sanctions', 'OUI→Oui, NON→Non',
                'low', 'Sanctions internationales — à valider')

    # === Anti-corruption ===
    if 'corruption' in t.lower() or 'corruption' in q:
        return ('GAP', None, 'OUI→Oui, NON→Non',
                'low', 'GAP : pas de variable diagnostic anti-corruption. Ajouter au schema ou flagger À_VALIDER')

    # === Fraude ===
    if 'fraude' in t.lower() or 'fraude' in q:
        return ('GAP', None, 'OUI→Oui, NON→Non',
                'low', 'GAP : pas de variable diagnostic prévention fraude. Ajouter au schema ou flagger À_VALIDER')

    # === Réclamations ===
    if 'réclamation' in t.lower() or 'réclamation' in q:
        if 'reporting' in q:
            return ('bloc_8', 'suivi_incidents', 'OUI→Oui, NON→Non',
                    'medium', 'Reporting réclamations')
        return ('bloc_8', 'identification_incidents', 'OUI→Oui, PARTIEL→Non',
                'medium', 'Politique réclamations')

    # === Archivage ===
    if 'archivage' in t.lower():
        if 'données de santé' in q or 'hds' in q.lower():
            return ('bloc_5', 'zones_aeras_securisees', 'OUI→Oui, NON→Non',
                    'medium', 'Archivage données santé')
        return ('GAP', None, 'OUI→Oui, NON→Non',
                'low', 'GAP : pas de variable diagnostic archivage. Ajouter au schema ou flagger À_VALIDER')

    # === Contrôle interne ===
    if 'contrôle interne' in t.lower() or 'cartographie' in q or 'procédures' in q and 'mode' in q:
        return ('bloc_1', 'controle_interne',
                'OUI→Oui, PARTIEL→Non',
                'high', 'Convention : PARTIEL → Non + commentaire')

    # === Habilitations / Séparation ===
    if 'habilitations' in t.lower() or 'séparation des tâches' in q or 'pouvoirs d\'engagement' in q:
        return ('bloc_4', 'acces_individualises', 'OUI→Oui, NON→Non',
                'medium', 'Habilitations/séparation des tâches')

    # === PCA / PCI ===
    if 'pca' in t.lower() or 'continuité' in t.lower() or 'continuité' in q:
        if 'informatique' in q or 'applications' in q or 'pci' in q.lower():
            return ('bloc_4', 'continuite_activite', 'OUI→Oui, NON→Non',
                    'high', 'PCI informatique')
        return ('bloc_1', 'pca', 'OUI→Oui, NON→Non',
                'high', 'PCA métier')
    if 'plan de réversibilité' in q:
        return ('bloc_6', 'controle_activite_deleguee', 'OUI→Oui, NON→Non',
                'low', 'Plan réversibilité subdélégataires')

    # === Sécurité SI / Exploitation ===
    if 'sécurité' in t.lower() and 'information' in t.lower():
        if 'politique' in q and 'sécurité' in q:
            return ('bloc_4', 'stockage_centralise', 'OUI→Oui, NON→Non',
                    'medium', 'Politique SSI — proxy stockage centralisé')
        if 'firewall' in q or 'vpn' in q:
            return ('bloc_4', 'securite_acces', 'OUI→Oui, PARTIEL→Non',
                    'medium', '')
        if 'intrusions physiques' in q or 'locaux' in q:
            return ('bloc_4', 'securite_acces', 'OUI→Oui, PARTIEL→Non',
                    'low', 'Sécurité physique locaux')
        return ('bloc_4', 'stockage_centralise', 'OUI→Oui, NON→Non',
                'low', 'Sécurité SI générique')

    if 'exploitation' in t.lower():
        if 'sauvegarde' in q:
            return ('bloc_4', 'sauvegardes', 'OUI→Oui, PARTIEL→Non, NON→Non',
                    'high', '')
        if 'séparation' in q and 'environnement' in q:
            return ('bloc_4', 'stockage_centralise', 'OUI→Oui, NON→Non',
                    'low', 'Séparation prod/hors-prod')
        if 'malveillant' in q or 'intrusions logiques' in q or 'attaques' in q:
            return ('bloc_4', 'securite_acces', 'OUI→Oui, PARTIEL→Non',
                    'medium', 'Protection malware')
        if 'vulnérabilité' in q or 'scan' in q:
            return ('bloc_4', 'securite_acces', 'OUI→Oui, PARTIEL→Non',
                    'low', 'Scans vulnérabilité')
        return ('bloc_4', 'securite_acces', 'OUI→Oui, PARTIEL→Non',
                'low', 'Sécurité exploitation')

    # === Cryptographie ===
    if 'cryptographie' in t.lower() or 'chiffrement' in q or 'chiffrez' in q:
        return ('bloc_4', 'securite_acces', 'OUI→Oui, PARTIEL→Non',
                'medium', 'Chiffrement')

    # === Communications ===
    if 'communications' in t.lower() or 'interconnexion' in q or 'cartographie' in q and 'système' in q:
        if 'cartographie' in q:
            return ('bloc_4', 'sous_traitants_tic_critiques', 'OUI→Oui, PARTIEL→Non',
                    'medium', 'Cartographie SI')
        return ('bloc_4', 'securite_acces', 'OUI→Oui, PARTIEL→Non',
                'low', 'Sécurité communications')

    # === Relations fournisseurs / TIC ===
    if 'fournisseurs' in t.lower() or 'sous-traitant' in q or 'prestation' in q and 'informatique' in q:
        if 'cloud' in q.lower():
            return ('bloc_4', 'sous_traitants_tic_critiques', 'OUI→Oui, PARTIEL→Non',
                    'medium', 'Cloud — sous-traitant TIC')
        if 'hors ue' in q or 'hors eee' in q:
            return ('bloc_1', 'activites_operationnelles_traitees_hors_france',
                    'OUI→Oui, NON→Non',
                    'medium', 'Sous-traitance hors UE')
        if 'sous-traitants critiques' in q or 'tic' in q.lower():
            return ('bloc_4', 'sous_traitants_tic_critiques', 'OUI→Oui, PARTIEL→Non',
                    'high', '')
        return ('bloc_4', 'sous_traitants_tic_critiques', 'OUI→Oui, PARTIEL→Non',
                'medium', 'Relations fournisseurs')

    # === Gestion incidents informatiques ===
    if 'incidents' in t.lower() and ('information' in t.lower() or 'informatique' in t.lower()):
        return ('bloc_8', 'identification_incidents',
                'OUI→Oui, PARTIEL→Non',
                'high', 'Incidents informatiques')

    # === Plan continuité informatique (PCI) ===
    if 'continuité informatique' in t.lower() or 'pci' in t.lower():
        return ('bloc_4', 'continuite_activite', 'OUI→Oui, NON→Non',
                'high', 'PCI')

    # === Contrôle d'accès ===
    if 'contrôle d\'accès' in t.lower() or 'authentification' in q:
        return ('bloc_4', 'acces_individualises', 'OUI→Oui, NON→Non',
                'high', '')

    # === CNIL / RGPD ===
    if 'rgpd' in t.lower() or 'cnil' in t.lower() or 'données personnelles' in q or 'données à caractère personnel' in q:
        if 'dpo' in q.lower() or 'délégué à la protection' in q.lower():
            return ('bloc_1', 'rgpd', 'OUI→Oui, PARTIEL→Non',
                    'medium', 'DPO — proxy RGPD')
        if 'registre' in q and 'traitement' in q:
            return ('bloc_1', 'rgpd', 'OUI→Oui, PARTIEL→Non',
                    'high', 'Registre traitements')
        if 'eee' in q or 'espace économique européen' in q:
            return ('bloc_1', 'activite_hors_eee', 'OUI→Non, NON→Oui',
                    'medium', 'Flux EEE — logique inversée')
        if 'formation' in q:
            return ('bloc_1', 'rgpd', 'OUI→Oui, PARTIEL→Non',
                    'medium', 'Formation RGPD — proxy')
        if 'violations' in q or 'incidents de sécurité' in q:
            return ('bloc_8', 'identification_incidents',
                    'OUI→Oui, PARTIEL→Non',
                    'medium', 'Violations données')
        if 'mentions légales' in q:
            return ('bloc_1', 'rgpd', 'OUI→Oui, PARTIEL→Non',
                    'low', 'Mentions légales')
        return ('bloc_1', 'rgpd', 'OUI→Oui, PARTIEL→Non',
                'medium', 'RGPD générique')

    # === Qualité des données ===
    if 'qualité des données' in t.lower():
        return ('GAP', None, 'OUI→Oui, NON→Non',
                'low', 'GAP : pas de variable diagnostic qualité données. Ajouter au schema ou flagger À_VALIDER')

    # === RSE ===
    if 'rse' in t.lower():
        return ('GAP', None, 'OUI→Oui, NON→Non',
                'low', 'GAP : pas de variable diagnostic RSE. Ajouter au schema ou flagger À_VALIDER')

    # === IA ===
    if 'intelligence artificielle' in t.lower() or 'règlement ia' in q:
        if 'subdélégataire' in q or 'subdélégué' in q:
            return ('bloc_6', 'ia_subdelegataire', 'CONFORME→Oui, NON→Non',
                    'high', '')
        return ('bloc_1', 'ia_conformite', 'OUI→Oui, NON→Non, N/A→N/A',
                'high', '')

    # === Données de santé (bloc 5) ===
    if 'santé' in t.lower() or 'médical' in t.lower() or 'médecin' in q or 'secret médical' in q or 'aeras' in q.lower() or 'données de santé' in q:
        if 'médecin conseil' in q:
            return ('bloc_5', 'traitement_sante', 'ACTIF→Oui, INACTIF→N/A',
                    'medium', 'Médecin conseil')
        if 'aeras' in q.lower() or 'zones' in q:
            return ('bloc_5', 'zones_aeras_securisees', 'OUI→Oui, NON→Non',
                    'high', '')
        if 'accès' in q and 'restreint' in q:
            return ('bloc_5', 'acces_restreint', 'OUI→Oui, NON→Non',
                    'high', '')
        if 'formation' in q or 'secret médical' in q:
            return ('bloc_5', 'acces_restreint', 'OUI→Oui, NON→Non',
                    'medium', 'Formation données santé')
        if 'flux' in q:
            return ('bloc_5', 'acces_restreint', 'OUI→Oui, NON→Non',
                    'medium', 'Flux données santé')
        if 'archives' in q or 'archivage' in q:
            return ('bloc_5', 'zones_aeras_securisees', 'OUI→Oui, NON→Non',
                    'medium', 'Archivage santé')
        if 'procédure' in q or 'dispositif' in q:
            return ('bloc_5', 'acces_restreint', 'OUI→Oui, NON→Non',
                    'low', 'Procédure santé')
        if 'contrôle' in q or 'revue' in q:
            return ('bloc_5', 'acces_restreint', 'OUI→Oui, NON→Non',
                    'low', 'Contrôles santé')
        return ('bloc_5', 'acces_restreint', 'OUI→Oui, NON→Non',
                'low', 'Santé générique')

    # === Subdélégation (bloc 6) ===
    if 'subdél' in t.lower() or 'subdél' in q:
        if 'convention' in q or 'contrat' in q:
            return ('bloc_6', 'contrat_subdelegation', 'OUI→Oui, NON→Non',
                    'high', '')
        if 'pilotage' in q or 'instances' in q or 'reporting' in q:
            return ('bloc_6', 'supervision', 'OUI→Oui, NON→Non',
                    'high', '')
        if 'contrôle' in q or 'audit' in q:
            return ('bloc_6', 'controle_activite_deleguee', 'OUI→Oui, NON→Non',
                    'high', '')
        if 'eee' in q.lower() or 'cct' in q.lower() or 'clauses contractuelles' in q:
            return ('bloc_6', 'contrat_subdelegation', 'OUI→Oui, NON→Non',
                    'medium', 'CCT subdél hors EEE')
        if 'cloud' in q.lower():
            return ('bloc_6', 'controle_activite_deleguee', 'OUI→Oui, NON→Non',
                    'medium', 'Cloud subdél')
        if 'ia' in q.lower() or 'intelligence' in q.lower():
            return ('bloc_6', 'ia_subdelegataire', 'CONFORME→Oui, NON→Non',
                    'high', '')
        if 'réversibilité' in q:
            return ('bloc_6', 'controle_activite_deleguee', 'OUI→Oui, NON→Non',
                    'medium', 'Plan réversibilité')
        if 'conformité' in q:
            return ('bloc_6', 'controle_activite_deleguee', 'OUI→Oui, NON→Non',
                    'medium', 'Suivi conformité subdél')
        return ('bloc_6', 'sous_delegation_active', 'ACTIF→Oui, INACTIF→N/A',
                'low', 'Subdélégation générique')

    # === Bloc 7 - Encaissement (feuille "Délégation d'encaissement-décai") ===
    # Cette feuille ne contient que des questions PJ déjà traitées plus haut

    # === RC Pro / Garantie financière ===
    if 'responsabilité civile' in t.lower() or 'rc pro' in q or 'garantie' in q and 'professionnelle' in q:
        return ('GAP', None, 'OUI→Oui, NON→Non',
                'medium', 'GAP : RC Pro pas couvert par diagnostic. Ajouter rc_pro_en_vigueur ?')

    if 'garantie financière' in t.lower() or 'caution bancaire' in t.lower():
        return ('bloc_7', 'encaissement_actif', 'ACTIF→Oui, INACTIF→N/A',
                'medium', 'Garantie financière liée à encaissement')

    # === Connaissance activités (distributeur) ===
    if 'connaissance' in t.lower() and 'activit' in t.lower():
        return ('INPUT_DIRECT', None, 'valeur brute',
                'medium', 'Question d\'activité — données directes (énumération/pourcentage)')

    # === Vente à distance ===
    if 'vente à distance' in t.lower() or 'vente à distance' in q:
        return ('bloc_2', 'vente_distance', 'CONFORME→Oui, NON→Non, N/A→N/A',
                'high', '')

    # === Légal et réglementaire ===
    if 'légal et réglementaire' in t.lower():
        if 'documentation' in q and 'précontractuelle' in q:
            return ('bloc_2', 'tracabilite_conseil', 'OUI→Oui, PARTIEL→Non',
                    'medium', 'Documentation précontractuelle')
        if 'rémunération' in q:
            return ('bloc_2', 'recommandation_personnalisee', 'OUI→Oui, NON→Non',
                    'low', 'Info rémunération')
        if 'honorabilité' in q:
            return ('GAP', None, 'OUI→Oui, NON→Non',
                    'low', 'GAP : honorabilité annuelle non couverte')
        return ('bloc_2', 'tracabilite_conseil', 'OUI→Oui, PARTIEL→Non',
                'low', 'Légal/réglementaire générique')

    # === Gouvernance et surveillance produit (POG) ===
    if 'gouvernance' in t.lower() and 'produit' in t.lower():
        if 'connaissance' in q or 'compétences' in q:
            return ('bloc_2', 'adequation_produit_client', 'OUI→Oui, PARTIEL→Non',
                    'medium', 'Connaissance produits / compétences')
        if 'alerte' in q or 'inadéquat' in q:
            return ('bloc_8', 'identification_incidents',
                    'OUI→Oui, PARTIEL→Non',
                    'medium', 'Alerte produit inadéquat')
        if 'conflits d\'intérêts' in q or 'conflits d\'intérêt' in q:
            return ('GAP', None, 'OUI→Oui, NON→Non',
                    'low', 'GAP : conflits d\'intérêts non couverts')
        if 'rémunération' in q or 'incitation' in q:
            return ('GAP', None, 'OUI→Oui, NON→Non',
                    'low', 'GAP : incitations rémunération non couvertes')
        return ('bloc_2', 'adequation_produit_client', 'OUI→Oui, PARTIEL→Non',
                'low', 'POG générique')

    # === Éthique des affaires (Entrée en relation) ===
    if 'éthique' in t.lower():
        return ('GAP', None, 'OUI→Oui, NON→Non',
                'medium', 'GAP : Éthique des affaires (condamnations, PPE dirigeants) non couverte par diagnostic')

    # === Connaissance générale intermédiaire (entrée en relation) ===
    if 'connaissance générale' in t.lower():
        if 'courtage à titre accessoire' in q or 'l511-1' in q.lower():
            return ('bloc_1', 'activite_accessoire', 'OUI→Oui, NON→Non',
                    'high', '')
        if 'autre activité' in q or 'autre statut' in q:
            return ('GAP', None, 'OUI→Oui, NON→Non',
                    'medium', 'GAP : autres statuts intermédiation non couverts')
        return ('INPUT_DIRECT', None, 'valeur brute',
                'medium', 'Connaissance générale — donnée directe')

    # === Fallback ===
    return ('UNKNOWN', None, 'OUI→Oui, NON→Non',
            'low', 'Non classifié automatiquement — à mapper manuellement')


def main():
    all_rows = []
    for sheet_name in SHEETS_WITH_QUESTIONS:
        df = pd.read_excel(SOURCE, sheet_name=sheet_name)
        df = df[df['Question'].notna()]
        for _, row in df.iterrows():
            ref = str(row.get('Référence', '')).strip()
            theme = str(row.get('Thème', '')).strip() if pd.notna(row.get('Thème')) else ''
            question = str(row.get('Question', '')).strip()
            type_reponse = str(row.get('Type de réponse', '')).strip() if pd.notna(row.get('Type de réponse')) else ''
            nb_pj = row.get('NB PJ Demandées')
            nb_pj = int(nb_pj) if pd.notna(nb_pj) else 0

            bloc, var, conv, conf, notes = classify(ref, theme, question, type_reponse)

            all_rows.append({
                'Référence': ref,
                'Feuille EDI': sheet_name,
                'Thème': theme,
                'Question': question[:300],  # tronquer pour lisibilité
                'Type de réponse': type_reponse,
                'NB PJ': nb_pj,
                'Bloc source': bloc,
                'Variable diagnostic': var or '',
                'Conversion': conv or '',
                'Confidence': conf,
                'Validation': '',
                'Notes': notes
            })

    # Création du workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping EDI"

    headers = list(all_rows[0].keys())
    ws.append(headers)

    # Styles
    header_fill = PatternFill('solid', start_color='1F4E78')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    high_fill = PatternFill('solid', start_color='C6EFCE')   # vert clair
    medium_fill = PatternFill('solid', start_color='FFF2CC') # jaune clair
    low_fill = PatternFill('solid', start_color='F8CBAD')    # orange clair
    gap_fill = PatternFill('solid', start_color='FFC7CE')    # rouge clair
    border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    # Style header
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Lignes de données
    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.font = Font(name='Calibri', size=10)
            cell.border = border

        # Coloration par confidence
        conf = row_data['Confidence']
        bloc = row_data['Bloc source']
        if bloc == 'GAP':
            fill = gap_fill
        elif bloc == 'UNKNOWN':
            fill = gap_fill
        elif conf == 'high':
            fill = high_fill
        elif conf == 'medium':
            fill = medium_fill
        else:
            fill = low_fill
        ws.cell(row=row_idx, column=headers.index('Confidence') + 1).fill = fill

    # Validation dropdown sur colonne "Validation"
    dv = DataValidation(type="list", formula1='"OK,À corriger,Validé,À discuter"', allow_blank=True)
    dv.add(f'K2:K{len(all_rows)+1}')
    ws.add_data_validation(dv)

    # Largeurs colonnes
    col_widths = {
        'A': 14, 'B': 22, 'C': 28, 'D': 60, 'E': 18,
        'F': 8, 'G': 16, 'H': 28, 'I': 28, 'J': 12,
        'K': 14, 'L': 36
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 36

    # Freeze panes
    ws.freeze_panes = 'D2'

    # === Feuille STATISTIQUES ===
    ws_stats = wb.create_sheet('Statistiques')
    ws_stats.append(['Métrique', 'Valeur'])
    ws_stats.cell(row=1, column=1).font = header_font
    ws_stats.cell(row=1, column=2).font = header_font
    ws_stats.cell(row=1, column=1).fill = header_fill
    ws_stats.cell(row=1, column=2).fill = header_fill

    total = len(all_rows)
    by_bloc = {}
    by_conf = {'high': 0, 'medium': 0, 'low': 0}
    for r in all_rows:
        b = r['Bloc source']
        by_bloc[b] = by_bloc.get(b, 0) + 1
        by_conf[r['Confidence']] += 1

    ws_stats.append(['Total questions', total])
    ws_stats.append(['', ''])
    ws_stats.append(['Confidence haute (mapping fiable)', by_conf['high']])
    ws_stats.append(['Confidence moyenne (à valider)', by_conf['medium']])
    ws_stats.append(['Confidence basse (à revoir)', by_conf['low']])
    ws_stats.append(['', ''])
    ws_stats.append(['--- Répartition par bloc source ---', ''])
    for b in sorted(by_bloc.keys()):
        ws_stats.append([b, by_bloc[b]])

    ws_stats.column_dimensions['A'].width = 40
    ws_stats.column_dimensions['B'].width = 12

    # === Feuille LÉGENDE ===
    ws_leg = wb.create_sheet('Légende')
    legend = [
        ['Codes Bloc source', 'Signification'],
        ['bloc_1', 'Socle réglementaire — variables atomiques bloc 1'],
        ['bloc_2', 'Distribution / Conseil — variables atomiques bloc 2'],
        ['bloc_3', 'Connaissance client — variables atomiques bloc 3'],
        ['bloc_4', 'Système d\'information — variables atomiques bloc 4'],
        ['bloc_5', 'Données sensibles santé — variables atomiques bloc 5'],
        ['bloc_6', 'Sous-délégation — variables atomiques bloc 6'],
        ['bloc_7', 'Encaissement — variables atomiques bloc 7'],
        ['bloc_8', 'Incidents / Pilotage — variables atomiques bloc 8'],
        ['PJ_DOC', 'Pièce justificative — réponse = présence du doc (Kbis, ORIAS, etc.)'],
        ['INPUT_DIRECT', 'Donnée brute à extraire (effectif, CA, %, énumération)'],
        ['GAP', 'Aucune variable diagnostic ne couvre — soit étendre le schema, soit flagger À_VALIDER'],
        ['UNKNOWN', 'Non classifié automatiquement — à mapper manuellement'],
        ['', ''],
        ['Codes Conversion', ''],
        ['OUI→Oui, NON→Non', 'Mapping direct'],
        ['OUI→Non, NON→Oui', 'Logique INVERSÉE (la question EDI inverse la sémantique)'],
        ['OUI→Oui, PARTIEL→Non', 'PARTIEL est traité conservativement comme Non + commentaire'],
        ['ACTIF→Oui, INACTIF→N/A', 'Pour les blocs neutralisables (5/6/7)'],
        ['valeur brute', 'Le diag fournit directement la valeur (nombre, texte)'],
        ['', ''],
        ['Couleurs Confidence', ''],
        ['Vert', 'Mapping haute confiance — validation rapide'],
        ['Jaune', 'Mapping confiance moyenne — vérification recommandée'],
        ['Orange', 'Mapping confiance faible — à revoir'],
        ['Rouge', 'GAP ou non classifié — décision métier requise'],
    ]
    for r in legend:
        ws_leg.append(r)
    ws_leg.cell(row=1, column=1).font = header_font
    ws_leg.cell(row=1, column=2).font = header_font
    ws_leg.cell(row=1, column=1).fill = header_fill
    ws_leg.cell(row=1, column=2).fill = header_fill
    ws_leg.column_dimensions['A'].width = 28
    ws_leg.column_dimensions['B'].width = 90

    wb.save(DEST)
    print(f"Saved {DEST}")
    print(f"Total questions: {total}")
    print(f"By confidence: {by_conf}")
    print(f"By bloc:")
    for b in sorted(by_bloc.keys()):
        print(f"  {b}: {by_bloc[b]}")


if __name__ == '__main__':
    main()
